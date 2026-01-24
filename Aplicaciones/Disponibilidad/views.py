from django.db import transaction
from django.db.models import Sum
from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import datetime
from django.utils import timezone
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer

from rest_framework import viewsets, status
from rest_framework.decorators import api_view, action
from rest_framework.response import Response

from .serializers import DisponibilidadSerializer
from .models import Disponibilidad, QRDisponibilidadUsado,Variedad

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated


from .models import Variedad
from .serializers import VariedadSerializer

from django.db.models.deletion import ProtectedError



def inicio(request):
   
    disponibilidades = Disponibilidad.objects.all()
    return render(request, 'disponibilidad.html', {
        'disponibilidades': disponibilidades
    })



def eliminar_disponibilidad(request, id):
    Disponibilidad.objects.get(id=id).delete()
    messages.success(request, "Disponibilidad eliminada correctamente")
    return redirect('dispo')


def procesar_edicion_disponibilidad(request):
    if request.method == "POST":
        try:
            _id = (request.POST.get("id") or "").strip()

            numero_mesa = request.POST.get("numero_mesa")
            variedad = (request.POST.get("variedad") or "").strip()
            medida = (request.POST.get("medida") or "").strip()
            stock = int(request.POST.get("stock") or 0)

            # fecha_entrada (datetime-local)
            fecha_entrada = None
            if request.POST.get("fecha_entrada"):
                fecha_entrada = datetime.strptime(
                    request.POST["fecha_entrada"], "%Y-%m-%dT%H:%M"
                )
                # opcional: hacer aware
                if timezone.is_naive(fecha_entrada):
                    fecha_entrada = timezone.make_aware(fecha_entrada, timezone.get_current_timezone())
            else:
                fecha_entrada = timezone.now()

            # fecha_salida
            fecha_salida = None
            if request.POST.get("fecha_salida"):
                fecha_salida = datetime.strptime(
                    request.POST["fecha_salida"], "%Y-%m-%dT%H:%M"
                )
                if timezone.is_naive(fecha_salida):
                    fecha_salida = timezone.make_aware(fecha_salida, timezone.get_current_timezone())

            # ==========================
            # ✅ UPSERT (EDITA o CREA)
            # ==========================
            if _id:
                d = Disponibilidad.objects.get(id=_id)  # edita
                d.numero_mesa = numero_mesa
                d.variedad = variedad
                d.medida = medida
                d.stock = stock
                d.fecha_entrada = fecha_entrada
                d.fecha_salida = fecha_salida
                d.save()

                msg = "Disponibilidad actualizada correctamente"
            else:
                # crea un registro nuevo (editar el 0)
                d = Disponibilidad.objects.create(
                    numero_mesa=numero_mesa,
                    variedad=variedad,
                    medida=medida,
                    stock=stock,
                    fecha_entrada=fecha_entrada,
                    fecha_salida=fecha_salida
                )

                msg = "Disponibilidad creada (se editó el 0) correctamente"

            # ==========================
            # ✅ WEBSOCKET igual que antes
            # ==========================
            async_to_sync(get_channel_layer().group_send)(
                "disponibilidad",
                {
                    "type": "nueva_disponibilidad",
                    "data": DisponibilidadSerializer(d).data
                }
            )

            messages.success(request, msg)

        except Exception as e:
            messages.error(request, f"Error: {e}")

        return redirect('dispo')



# =========================
# API REST – VIEWSET
# =========================

class DisponibilidadViewSet(viewsets.ModelViewSet):
    queryset = Disponibilidad.objects.all().order_by('-fecha_entrada')
    serializer_class = DisponibilidadSerializer

    @action(detail=False, methods=['get'])
    def activos(self, request):
        qs = Disponibilidad.objects.filter(fecha_salida__isnull=True)
        return Response(self.get_serializer(qs, many=True).data)

    @action(detail=False, methods=['get'])
    def por_mesa(self, request):
        mesa = request.query_params.get("mesa")
        if not mesa:
            return Response({"error": "Parámetro mesa requerido"}, status=400)
        qs = Disponibilidad.objects.filter(numero_mesa=mesa)
        return Response(self.get_serializer(qs, many=True).data)


# =========================
# API MANUAL
# =========================
@api_view(['GET', 'POST'])
def api_disponibilidad_list(request):
    print("👤 user:", request.user, "auth:", request.user.is_authenticated)
    print("🍪 cookies:", request.COOKIES)
    print("📌 session keys:", list(request.session.keys()))

    if request.method == 'GET':
        ordenar = request.query_params.get("ordenar")
        fecha = request.query_params.get("fecha")
        desde = request.query_params.get("desde")
        hasta = request.query_params.get("hasta")
        reciente = request.query_params.get("reciente")
        
        qs = Disponibilidad.objects.all()
        
        if fecha:
            qs = qs.filter(fecha_entrada__date=fecha)

        if desde and hasta:
            qs = qs.filter(fecha_entrada__date__range=[desde, hasta])

        campos = {
            "mesa": "numero_mesa",
            "variedad": "variedad",
            "medida": "medida",
            "fecha": "fecha_entrada"
        }

        if ordenar in campos:
            campo = campos[ordenar]
            if reciente == "true":
                campo = "-" + campo
            qs = qs.order_by(campo)
        
        return Response(DisponibilidadSerializer(qs, many=True).data)

    elif request.method == 'POST':

        data = request.data
        codigo = data.get("qr_id")
        mesa = data.get("numero_mesa")
        variedad = data.get("variedad")
        medida = data.get("medida")

        if not codigo or not mesa or not variedad or not medida:
            return Response({"error": "Datos incompletos"}, status=status.HTTP_400_BAD_REQUEST)


        if QRDisponibilidadUsado.objects.filter(qr_id=codigo).exists():
            return Response(
                {"error": "Este QR ya fue utilizado en Disponibilidad"},
                status=status.HTTP_409_CONFLICT
            )

        # Guardar QR para siempre
        QRDisponibilidadUsado.objects.create(qr_id=codigo)

        hoy = timezone.localdate()

        existente = Disponibilidad.objects.filter(
            numero_mesa=mesa,
            variedad=variedad,
            medida=medida,
            fecha_entrada__date=hoy
        ).first()

        if existente:
            existente.stock += 1
            existente.save()

            async_to_sync(get_channel_layer().group_send)(
                "disponibilidad",
                {
                    "type": "nueva_disponibilidad",
                    "data": DisponibilidadSerializer(existente).data
                }
            )
            return Response(DisponibilidadSerializer(existente).data, status=200)

        nuevo = Disponibilidad.objects.create(
            numero_mesa=mesa,
            variedad=variedad,
            medida=medida,
            stock=1,
            fecha_entrada=timezone.now()
        )

        async_to_sync(get_channel_layer().group_send)(
            "disponibilidad",
            {
                "type": "nueva_disponibilidad",
                "data": DisponibilidadSerializer(nuevo).data
            }
        )

        return Response(DisponibilidadSerializer(nuevo).data, status=201)


@api_view(['GET', 'PUT', 'DELETE'])
def api_disponibilidad_detail(request, pk):
    try:
        disponibilidad = Disponibilidad.objects.get(pk=pk)
    except Disponibilidad.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = DisponibilidadSerializer(disponibilidad)
        return Response(serializer.data)

    elif request.method == 'PUT':
        serializer = DisponibilidadSerializer(disponibilidad, data=request.data)
        if serializer.is_valid():
            serializer.save()

            async_to_sync(get_channel_layer().group_send)(
                "disponibilidad",
                {
                    "type": "nueva_disponibilidad",
                    "data": serializer.data
                }
            )

            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        disponibilidad.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

@api_view(['GET'])
def api_disponibilidad_stats(request):
    return Response({
        "total_registros": Disponibilidad.objects.count(),
        "registros_activos": Disponibilidad.objects.filter(fecha_salida__isnull=True).count(),
        "stock_total": Disponibilidad.objects.aggregate(Sum('stock'))['stock__sum'] or 0,
        "mesas_activas": Disponibilidad.objects.filter(fecha_salida__isnull=True)
                            .values('numero_mesa')
                            .distinct()
                            .count()
    })
#API PARA LA DISPONIBILIDAD QUE SALE
from .models import Disponibilidad, QRDisponibilidadSalidaUsado

@api_view(['POST'])
def api_disponibilidad_salida(request):
    data = request.data
    codigo = data.get("qr_id")
    variedad = data.get("variedad")
    medida = data.get("medida")

    if not codigo or not variedad or not medida:
        return Response(
            {"error": "Datos incompletos: qr_id, variedad y medida son obligatorios"},
            status=status.HTTP_400_BAD_REQUEST
        )

    # ⛔ Si ya se restó este QR una vez, NO permitir otra vez
    if QRDisponibilidadSalidaUsado.objects.filter(qr_id=codigo).exists():
        return Response(
            {"error": "Este QR ya fue utilizado en SALIDA (ya se restó una vez)"},
            status=status.HTTP_409_CONFLICT
        )

    channel_layer = get_channel_layer()

    with transaction.atomic():
        dispo = (Disponibilidad.objects
                 .select_for_update()
                 .filter(
                     fecha_salida__isnull=True,
                     variedad=variedad,
                     medida=medida,
                     stock__gt=0
                 )
                 .order_by('fecha_entrada', 'id')
                 .first())

        if not dispo:
            return Response(
                {"error": "No hay stock disponible para esa variedad y medida"},
                status=status.HTTP_409_CONFLICT
            )

        # Registrar QR como ya restado
        QRDisponibilidadSalidaUsado.objects.create(qr_id=codigo)

        # Restar 1
        dispo.stock -= 1

        # Si llega a 0, marca fecha_salida (opcional)
        if dispo.stock == 0:
            dispo.fecha_salida = timezone.now()

        dispo.save()

    # Notificar por websocket
    async_to_sync(channel_layer.group_send)(
        "disponibilidad",
        {
            "type": "nueva_disponibilidad",
            "data": DisponibilidadSerializer(dispo).data
        }
    )

    return Response(DisponibilidadSerializer(dispo).data, status=status.HTTP_200_OK)
################################
#API VARIEDAD#
################################
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def variedades_api(request):
    if request.method == "GET":
        qs = Variedad.objects.all().order_by("nombre")
        return Response(VariedadSerializer(qs, many=True).data)

    nombre = (request.data.get("nombre") or "").strip()
    if not nombre:
        return Response({"detail": "El nombre es obligatorio."}, status=status.HTTP_400_BAD_REQUEST)

    existe = Variedad.objects.filter(nombre__iexact=nombre).first()
    if existe:
        return Response(VariedadSerializer(existe).data, status=status.HTTP_200_OK)

    nueva = Variedad.objects.create(nombre=nombre)
    return Response(VariedadSerializer(nueva).data, status=status.HTTP_201_CREATED)

from openpyxl import load_workbook
from io import BytesIO

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def variedades_excel_api(request):
    """
    Recibe un archivo Excel con variedades.
    Formatos aceptados:
    - Columna con encabezado: 'variedad'
    - O primera columna sin encabezado
    """
    file = request.FILES.get("file")
    if not file:
        return Response({"detail": "Debes enviar un archivo en 'file'."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        wb = load_workbook(filename=BytesIO(file.read()), data_only=True)
        ws = wb.active

        # leer filas
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({"detail": "El Excel está vacío."}, status=status.HTTP_400_BAD_REQUEST)

        # detectar header
        header = [str(x).strip().lower() if x is not None else "" for x in rows[0]]
        idx = None
        if "variedad" in header:
            idx = header.index("variedad")
            data_rows = rows[1:]
        else:
            idx = 0
            data_rows = rows

        nombres = []
        for r in data_rows:
            if not r or len(r) <= idx:
                continue
            val = r[idx]
            if val is None:
                continue
            nombre = str(val).strip()
            if nombre:
                nombres.append(nombre)

        # quitar duplicados (ignorando mayúsculas)
        unicos = []
        seen = set()
        for n in nombres:
            k = n.lower()
            if k not in seen:
                seen.add(k)
                unicos.append(n)

        creadas = 0
        existentes = 0

        for n in unicos:
            if Variedad.objects.filter(nombre__iexact=n).exists():
                existentes += 1
            else:
                Variedad.objects.create(nombre=n)
                creadas += 1

        return Response({
            "detail": "ok",
            "creadas": creadas,
            "existentes": existentes,
            "total": len(unicos)
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({"detail": f"Error leyendo Excel: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
class VariedadViewSet(viewsets.ModelViewSet):
    
    queryset = Variedad.objects.all().order_by("nombre")
    serializer_class = VariedadSerializer

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            self.perform_destroy(instance)
            return Response(status=status.HTTP_204_NO_CONTENT)
        except ProtectedError:
            return Response(
                {"detail": "No puedes borrar esta variedad porque tiene stock/movimientos."},
                status=status.HTTP_409_CONFLICT
            )
@require_http_methods(["DELETE"])
def variedad_detail_api(request, pk):
    try:
        variedad = Variedad.objects.get(pk=pk)
    except Variedad.DoesNotExist:
        return JsonResponse({"detail": "La variedad no existe o ya fue eliminada."}, status=404)

    # ✅ bloquear si hay disponibilidad relacionada (ajusta el campo según tu modelo)
    # Caso 1: disponibilidad guarda variedad como texto -> comparar por nombre
    tiene = Disponibilidad.objects.filter(variedad__iexact=variedad.nombre).exists()

    # Caso 2 (si disponibilidad tiene FK): Disponibilidad.objects.filter(variedad_id=pk).exists()

    if tiene:
        return JsonResponse(
            {"detail": "No puedes borrar esta variedad porque tiene stock/movimientos."},
            status=409
        )

    variedad.delete()
    return JsonResponse({}, status=204)